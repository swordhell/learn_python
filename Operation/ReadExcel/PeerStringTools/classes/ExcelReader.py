# pip3 install openpyxl
from openpyxl import load_workbook
import logging

logger = logging.getLogger('ExcelReader')
class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename=self.file_path)
        logger.debug(f"Loaded workbook from {self.file_path}")
        
    def read_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        records = []
        for row in sheet.iter_rows(values_only=True):
            records.append(row)
        return records

# Example usage:
if __name__ == "__main__":
    file_path = "example.xlsx"  # Replace with your Excel file path
    reader = ExcelReader(file_path)
    sheet_name = "Sheet1"  # Replace with the name of the sheet you want to read
    records = reader.read_sheet(sheet_name)
    for record in records:
        print(record)
