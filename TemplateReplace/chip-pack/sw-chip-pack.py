# pip install openpyxl
# pip install jinja2
#
# 生成配置信息的逻辑
#

from jinja2 import Environment, PackageLoader
from openpyxl import load_workbook
import codecs

filename=r'./res/数玩-资源包合约 资产合约 上新.xlsx'
sheet_name=r'卡片信息'
outputfilename=r'./output/card.ts'

wb=load_workbook(filename = filename,data_only=True)
ws1=wb.get_sheet_by_name(sheet_name)

allNode=[]

for i in range(4,ws1.max_row+1):
    AssetID=ws1['A{}'.format(i)].value
    From=ws1['B{}'.format(i)].value
    To=ws1['C{}'.format(i)].value
    Name=ws1['G{}'.format(i)].value
    Rarity=ws1['M{}'.format(i)].value
    ShowName=ws1['N{}'.format(i)].value
    Description=ws1['P{}'.format(i)].value
    Image=ws1['S{}'.format(i)].value
    if AssetID == None:
        break;
    node = {}
    node['AssetID']=AssetID
    node['From']=From
    node['To']=To
    node['Name']=Name
    node['Rarity']=Rarity
    node['ShowName']=ShowName
    node['Description']=Description
    node['Image']=Image

    allNode.append(node)

env = Environment(loader=PackageLoader('sw-chip-pack', 'templates'))
template = env.get_template('card.tpl')

outputstr = template.render(allNode=allNode)

with codecs.open(outputfilename,"w",encoding='utf-8') as fd:
    fd.write(outputstr)